import openpyxl
import json

days = ['ПОНЕДЕЛЬНИК', "ВТОРНИК", 'СРЕДА', 'ЧЕТВЕРГ', 'ПЯТНИЦА', 'СУББОТА']
type_of_activity = ['ЛК', 'ПР', 'СР']
auditorium = ['ауд.', 'физ.', 'комп.']
group_temp = ['КМБО-02-23', 'КМБО-05-23', 'КМБО-02-22', 'КМБО-05-22', 'КМБО-05-21', 'КМБО-02-21', 'КМБО-02-20',
              'КМБО-05-20', "КММО-01-23", "КММО-01-22", "КММО-02-22"]
is_none = [None, None]
const_include_days = 5
const_not_include_days = 4
day_week = 14
min_row = 4
max_row = 87
min_col_main = 1
max_col_main = 5

'''
    Получаем список, который сожержит все дни недели, нач. и оконч. пар, номер пары и четность недели
'''


def get_main(book):
    sheet = book.active
    count_day = 0
    result = list()
    # Проходим по первым 5 колонкам файла, чтобы получить дни недели, нач. и оконч. пар, номер пары и четность недели
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col_main, max_col=max_col_main, values_only=True):
        temp = ''
        for cel in row:
            if cel is not None:
                temp += cel + ' '
        list_lesson = temp.split()
        count_day += 1
        result.append(list_lesson)
    return result


'''
    Получаем список нужных нам групп, в котором содержатся:
    1. Индексы начала (откуда считывать)
    2. Индексы конца (до куда считывать)
    3. Названия групп

    Возвращает список с группами и выше перечисленными параметрами 
'''


def get_groups(book):
    sheet = book.active
    groups = list()
    count_group_id = 0
    # Проходим по всем столбцам второй строки, и ищем нужные нам группы, и индексы для их дальнейшего считывания
    for row in sheet.iter_rows(min_row=2, max_row=2, min_col=2, max_col=sheet.max_column, values_only=True):
        for cel in row:
            count_group_id += 1
            if cel in group_temp:
                temp = list()
                temp.append(count_group_id + 1)
                temp.append(count_group_id + 4)
                temp.append(cel)

                groups.append(temp)
    return groups


'''
    Обрабатываем курс
'''


def parse(sheet, groups, kurs_work, book):
    main_info = get_main(book)  # Получаем базовую информацию
    # Проходим по всем группам, которые нужны нам
    for el in groups:
        count_day = 0
        kurs_work[el[2]] = list()
        group_name = el[2]
        id_begin = el[0]
        id_end = el[1]

        temp_day = ''

        for row, el_main_info in zip(sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=id_begin, max_col=id_end),
                                     main_info):
            temp_str = ''
            for cel in row:
                if cel.value is not None:
                    temp_str += cel.value + ' '

            list_lesson = temp_str.split()
            list_lesson = el_main_info + list_lesson  # Здесь объеденяем базовую информацию, и переменную (сама пара и различные её атрибуты)
            dict_lesson = dict()

            if count_day == 0:  # Если list_lesson содержит День недели, то обновляем temp_day
                for el in days:
                    if el in list_lesson:
                        temp_day = el

            dict_lesson["День недели"] = temp_day

            if count_day == 0:  # Если list_lesson содержит День недели
                del list_lesson[0]

            dict_lesson["Номер пары"] = list_lesson[0]
            dict_lesson["Нач. занятия"] = list_lesson[1]
            dict_lesson["Оконч. занятия"] = list_lesson[2]
            dict_lesson["Неделя занятия"] = list_lesson[3]

            count_day += 1

            if count_day == day_week:  # Обнуляем, если дошли до конца дня недели
                count_day = 0

            if len(list_lesson) == const_not_include_days:  # Если пары нет, то добавляем пустые строки
                dict_lesson["Дисциплина занятия"] = ""
                dict_lesson["Вид занятия"] = ""
                dict_lesson["ФИО преподавателя"] = ""
                dict_lesson["Номер аудитории"] = ""
            else:
                indices_name = [i for i, x in enumerate(list_lesson) if
                                x in type_of_activity]  # Индексы где встречаются типы занятия
                indices_values_name = [x for i, x in enumerate(list_lesson) if
                                       x in type_of_activity]  # Значения где встречаются  типы занятия
                indices_auditorium = [i for i, x in enumerate(list_lesson) if
                                      x in auditorium]  # Индексы где встречаются типы аудиторий

                if len(indices_name) == 1:  # Если одно занятия в строке

                    discipline_lesson = list_lesson[4:indices_name[0]]
                    discipline_lesson_str = ' '.join(discipline_lesson)
                    dict_lesson["Дисциплина занятия"] = discipline_lesson_str

                    if not indices_values_name:
                        dict_lesson["Вид занятия"] = ""
                    else:
                        dict_lesson["Вид занятия"] = indices_values_name[0]

                    if not indices_auditorium:
                        dict_lesson["ФИО преподавателя"] = ""
                        dict_lesson["Номер аудитории"] = ""
                    else:
                        full_name = list_lesson[(indices_name[0] + 1):indices_auditorium[0]]
                        dict_lesson["ФИО преподавателя"] = ' '.join(full_name)

                        audience_number = list_lesson[indices_auditorium[0]:]
                        audience_number_str = ' '.join(audience_number)

                        dict_lesson["Номер аудитории"] = audience_number_str

                elif len(indices_name) == 2:  # Если два занятия в строке

                    discipline_lesson = list_lesson[4:indices_name[0]]
                    discipline_lesson_str = ' '.join(discipline_lesson)
                    dict_lesson["Дисциплина занятия"] = discipline_lesson_str

                    if not indices_values_name:
                        dict_lesson["Вид занятия"] = ""
                    else:
                        dict_lesson["Вид занятия"] = ', '.join(indices_values_name)

                    if not indices_auditorium:
                        dict_lesson["ФИО преподавателя"] = ""
                        dict_lesson["Номер аудитории"] = ""
                    else:
                        full_name = list_lesson[(indices_name[1] + 1):indices_auditorium[0]]
                        dict_lesson["ФИО преподавателя"] = ' '.join(full_name)

                        audience_number = list_lesson[indices_auditorium[0]:]
                        audience_number_str = ' '.join(audience_number)
                        dict_lesson["Номер аудитории"] = audience_number_str

            kurs_work[group_name].append(dict_lesson)  # Добавляем пару


'''
    Считывание и запись в эксель
'''


def parse_exel(kurs, book):
    sheet = book.active

    groups = get_groups(book)

    kurs_work = dict()

    parse(sheet, groups, kurs_work, book)

    with open(f"kurs_work-{kurs}.json", "w", encoding='utf-8') as outfile:
        json.dump(kurs_work, outfile, ensure_ascii=False, indent=4)


if __name__ == '__main__':
    book = openpyxl.load_workbook("1-kurs.xlsx", read_only=True)
    parse_exel(1, book)
    book = openpyxl.load_workbook("2-kurs.xlsx", read_only=True)
    parse_exel(2, book)
    book = openpyxl.load_workbook("3-kurs.xlsx", read_only=True)
    parse_exel(3, book)
    book = openpyxl.load_workbook("4-kurs.xlsx", read_only=True)
    parse_exel(4, book)
    book = openpyxl.load_workbook("mag-1-kurs.xlsx", read_only=True)
    parse_exel(5, book)
    book = openpyxl.load_workbook("mag-2-kurs.xlsx", read_only=True)
    parse_exel(6, book)
