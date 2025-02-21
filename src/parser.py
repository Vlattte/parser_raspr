"""Парсер расписания с сайта vega.mirea.ru"""

# std
from os import getenv
from os import path

from copy import deepcopy

import re

# datetime
from datetime import datetime
from datetime import time
from datetime import timedelta

# usefull
from openpyxl import load_workbook
from openpyxl.styles.colors import Color

from progress.bar import PixelBar
from dotenv import load_dotenv

# custom
from src import utils
from src.db_class import Database

from src.structs import CmdParams
from src.structs import ListData
from src.structs import CellColors

# from src.structs import GroupParts
from src.structs import LessonParts
from src.structs import WeeksParts


WEEKS = 17
AUTUMN_SUBSTR = "осен"
SPRING_SUBSTR = "весен"
WINTER_SUBSTR = "зимн"
SUMMER_SUBSTR = "летн"

# TODO добавить: I,IIн - все недели, просто не смерджено
#                6-7п. - на 6 и 7 парах
# TODO если ошиблись в типе файла ругаться и выбрать нужный тип


class VegaRaspParser:
    """Парсер excel расписания"""

    def __init__(self, cmd_params: CmdParams = None):
        # очищаем ли базу
        pre_clear = False
        overwrite_day_start = None
        overwrite_day_end = None
        if cmd_params is None:
            load_dotenv()
            # доп флаги
            pre_clear = getenv("PRE_CLEAR")
            pre_clear = bool(int(pre_clear))
            self.is_debug = getenv("IS_DEBUG")
            # файлы для парсинга
            self.default_filename = getenv("DEFAULT_FILENAME")
            self.session_filename = getenv("SESSION_FILENAME")
            # дата начала и конца семестра
            self.start_date = getenv("START_DATE")
            self.end_date = getenv("END_DATE")
            # период для перезаписи расписания
            overwrite_day_start = getenv("OVERWRITE_DAY_START")
            overwrite_day_end = getenv("OVERWRITE_DAY_END")
            # первый день обучения магистров
            self.magic_first_stud_date = getenv("MAGIC_START_DATE")
        else:
            pre_clear = cmd_params.pre_clear
            # файлы для парсинга
            if cmd_params.is_sem:
                self.default_filename = cmd_params.filename
            else:
                self.session_filename = cmd_params.filename
            # дата начала и конца семестра
            self.start_date = cmd_params.start_date
            self.end_date = cmd_params.end_date
            # период для перезаписи расписания
            overwrite_day_start = cmd_params.overwrite_date_start
            overwrite_day_end = cmd_params.overwrite_date_end
            # первый день обучения магистров
            self.magic_first_stud_date = cmd_params.magic_start_date

        self.is_default = False
        self.is_session = False
        self.remember_existing_filenames()

        self.first_day = None
        self.last_day = None
        self.fill_dates(overwrite_day_start, overwrite_day_end)

        # общие позиции колонок и строк
        self.group_row = 2
        self.weekday_col = 1
        self.para_col = 2

        # позиция ячейки с версией
        self.version_col = 1
        self.version = 0

        # код семестра расписания
        self.semcode = 192000

        # Год начала и конца обучения
        self.start_year = datetime.now().year
        self.end_year = self.start_year + 1

        # инициализация класса работы с БД
        self.db = Database(pre_clear)

    def remember_existing_filenames(self):
        """Выбран ли хотя бы один файл, если выбран, запомнить какой(ие)"""
        if self.default_filename is None and self.session_filename is None:
            raise ValueError("Не выбран ни один файл для парсинга")

        if self.default_filename is not None:
            if not path.exists(self.default_filename):
                raise ValueError(
                    f"Выбран несуществующий файл файл расписания семестра: {self.default_filename}"
                )
            self.is_default = True

        if self.session_filename is not None:
            if not path.exists(self.session_filename):
                raise ValueError(
                    f"Выбран несуществующий файл расписания сессии: {self.session_filename}"
                )
            self.is_session = True

    def fill_dates(self, overwrite_day_start, overwrite_day_end):
        """Проверка и заполнение нужных дат"""
        if self.start_date is None or self.end_date is None:
            raise ValueError("Добавьте даты начала и конца семестра")

        # первый день обучения для магистров (не задан -> будет как у всех)
        if self.magic_first_stud_date is None:
            self.magic_first_stud_date = self.start_date
        self.magic_first_stud_date = datetime.strptime(
            self.magic_first_stud_date, "%Y-%m-%d"
        ).date()

        if overwrite_day_start is None or overwrite_day_end is None:
            overwrite_day_start = self.start_date
            overwrite_day_end = self.end_date
        # день начала и конца записи в БД, остальное не трогаем
        self.first_day = datetime.strptime(overwrite_day_start, "%Y-%m-%d").date()
        self.last_day = datetime.strptime(overwrite_day_end, "%Y-%m-%d").date()

        # логгируем, что заполнили дни
        days_filles_msg = (
            f"<--Таблица sc_rasp18_days заполнена c {overwrite_day_start} " +
            f"по {overwrite_day_end}"
        )
        self.log_cmd_message(days_filles_msg)

    def change_to_session_params(self):
        """Меняет некоторые стандартные параметры на параметры расписания сессии"""
        self.version_col = 2
        self.group_row = 4

    def change_to_dafault_params(self):
        """Возвращает параметры к дефолтным"""
        # позиция ячейки с версией
        self.version_col = 1
        self.group_row = 2

    def parse(self):
        """Парсинг исходя из параметров .env"""
        self.db.set_conn()

        # заполняем кафедры
        self.db.fill_departments()

        # заполняем типы пар
        self.db.fill_worktypes()

        # если выбрали файл расписания семетра
        if self.is_default:
            self.parse_semestr()
        # если выбрали файл расписания сессии
        if self.is_session:
            self.parse_exam()

        self.db.close_conn()

    def log_cmd_message(self, msg: str):
        """Дебаг сообщения в консоль"""
        if not self.is_debug:
            return

        print(msg)

    def parse_semestr(self):
        """Парсинг расписания семестровых пар"""
        self.log_cmd_message(f"<--PARSING SEMESTR: {self.default_filename}-->")

        self.change_to_dafault_params()
        self.parse_excel_file(self.default_filename, default_rasp=True)

        self.log_cmd_message(f"<--RASP PARSED: {self.default_filename}-->")

    def parse_exam(self):
        """Парсинг расписания экзаменов"""
        self.log_cmd_message(f"<--PARSING SESSION: {self.session_filename}-->")

        self.change_to_session_params()
        self.parse_excel_file(self.session_filename, default_rasp=False)

        self.log_cmd_message(f"<--SESSION PARSED: {self.session_filename}-->")

    def parse_excel_file(self, file_name, default_rasp=True):
        """Парсим данные"""
        ws = load_workbook(filename=file_name).active
        # ------------------Определяем границы расписания------------------
        min_col = ws.min_column
        max_col = ws.max_column
        max_row = utils.get_max_row(ws)
        # -----------------------------------------------------------------
        # вытаскиваем все нужное из заголовка
        self.fill_rasp_title_parts(max_row, ws)
        # заполняем дни таблицы rasp18_days
        self.db.fill_rasp18_for_period(self.semcode, self.start_date, self.end_date)
        # очистка данных расписания в период между заданными в конфиге датами
        self.db.clear_rasp_data_between_weeks(
            semcode=self.semcode,
            is_semestr=default_rasp,
            start_date=self.first_day,
            end_date=self.last_day,
        )
        # ---------------Парсим расписание в цикле по группам--------------
        merged_cells = ws.merged_cells
        for col in range(min_col, max_col + 1):
            group_cell = ws.cell(self.group_row, col).value
            # если заголовок столбца пустой, пропускаем
            if not group_cell:
                continue
            if default_rasp:
                self.parse_default_col(col, ws, max_row, group_cell, merged_cells)
            else:
                self.parse_exam_col(col, ws, max_row, group_cell)

    def parse_default_col(self, col, ws, max_row, group_cell, merged_cells):
        """Разбор колонки группы"""
        group_parts = utils.get_group_parts(group_cell)
        group_name = group_parts.name
        group_id = self.db.set_group(group_name)

        order = 1
        prev_weekday = "ПН"

        # проходимся по текущему столбцу по всем строкам
        row = self.group_row + 1

        progress_bar = None
        if self.is_debug:
            progress_bar = PixelBar(
                group_name, max=max_row - row, suffix="%(percent)d%%"
            )
            progress_bar.check_tty = False
            progress_bar.start()

        # Магистерская группа или нет
        is_magic = utils.is_magic_group(group_name)

        while row < max_row:
            if self.is_debug:
                progress_bar.next()

            weekday = ws.cell(row, self.weekday_col).value
            # если новый день недели
            if prev_weekday != weekday and weekday:
                prev_weekday = weekday
            weekday = prev_weekday

            # вытаскиваем данные по текущей паре
            cur_order = ws.cell(row, self.para_col).value
            lesson_cell = ws.cell(row, col).value
            teacher = ws.cell(row, col + 1).value
            room = ws.cell(row, col + 2).value

            # если пустая клетка, идем дальше
            if lesson_cell is None:
                row += 1
                # если строка нечетная, то номер пары + 1
                if row % 2 == 0:
                    order += 1
                continue

            # если в строке нет номера пары
            if cur_order is not None:
                order = cur_order

            # является ли длинной парой (НИР, ВОЕНКА, условно на весь день)
            coord = ws.cell(row, col).coordinate
            lesson_count = utils.get_lesson_count(merged_cells, coord)

            # по цвету определяем пренадлежность к кафедре
            cur_color = ws.cell(row, col).fill.start_color.rgb
            department_id = self.get_dep_id(cel_color=cur_color)

            # разделение ячейки названия пары на части
            lesson_parts = self.get_lesson_parts(lesson_cell, is_magic)

            # если подгруппа ВЕГИ или ВМ, то переписываем
            if group_parts.sub_group != -1:
                lesson_parts.sub_group = group_parts.sub_group

            # если элементы lesson_parts массивы, то для каждго из них делем записи
            complex_len = 1
            time_parts = utils.get_time_from_lesson(lesson_cell)

            # если сложное название дисциплины (с временем и неделями)
            if len(time_parts) > 0:
                complex_len = len(time_parts)

            # заполнение БД новыми данными
            for comp_record in range(complex_len):
                if len(time_parts) > 0:
                    t = time_parts[comp_record]
                    lesson_parts.timestart = t.timestart
                    lesson_parts.timeend = t.timeend
                    if t.weeks is not None:
                        week_parts = self.get_weeks_parts(t.weeks, is_magic)

                        lesson_parts.parity = week_parts.parity
                        lesson_parts.weeks_list = week_parts.weeks_list
                        lesson_parts.weeks_text = week_parts.weeks_text

                    lesson_count = utils.get_lesson_count_str(
                        lesson_parts.timestart, lesson_parts.timeend
                    )

                for pair_num in range(order, order + lesson_count):
                    self.fill_group_day_db(
                        lesson_parts,
                        teacher,
                        pair_num,
                        weekday,
                        group_id,
                        room,
                        department_id,
                        is_magic,
                    )

                    # смещение, если пара длится несколько пар
                    if lesson_parts.timestart is not None:
                        lesson_parts.timestart = utils.get_time_by_order(
                            pair_num + 1
                        )

            # переходим к новой паре
            row += 1

        if self.is_debug:
            progress_bar.finish()

    def parse_exam_col(self, col, ws, max_row, group_cell):
        """Разбор колонки расписания экзаменов"""
        # убираем курс из названия группы

        # Делит ячейку группы на части: 1 курс\nКМБО-47-25
        _, group_name = group_cell.split("\n")

        # заполняем БД данными по группе
        group_id = self.db.set_group(group_name)

        prev_date_cell = None
        row = self.group_row + 2

        # progress_bar = PixelBar(group_name, max=max_row - row, suffix="%(percent)d%%")
        # progress_bar.check_tty = False
        # progress_bar.start()

        # проходимся по текущему столбцу по всем строкам
        prev_weekday = "ПН"
        while row < max_row:
            date_cell = ws.cell(row, self.para_col).value
            # если следующий день
            prev_date_cell, date_cell = utils.swap_with_prev_value(
                prev_date_cell, date_cell
            )
            exam_date, weekday = date_cell.split("\n")
            if prev_weekday != weekday:
                prev_weekday = weekday

            exam_type = ws.cell(row, col + 1).value
            lesson_name = ws.cell(row + 1, col).value
            teacher = ws.cell(row + 2, col).value
            room = ws.cell(row + 2, col + 1).value

            # время начала экзамена
            time_start = ws.cell(row, col).value
            time_end = time_start

            if time_start is None:
                # если только эта строка пустая, то это разделитель
                is_splitter_row = utils.is_hsplitter(ws, row)
                if is_splitter_row:
                    # progress_bar.next()
                    row += 1
                    continue

                # если все ячейки данных по экзамену пустые
                if lesson_name is None and teacher is None:
                    # progress_bar.next(3)
                    row += 3
                    continue

                # если предмет не пустой, то ставим на 4 пары
                if teacher is None:
                    time_start = time(hour=9, minute=0)
                    time_end = time(hour=15, minute=50)

            # ставим длительность экзамена 1.5 часа (2 академ часа)
            else:
                time_end = utils.time_in_90_minutes(time_start)

            cur_order = utils.get_order_by_time(time_start)

            # Заполение таблиц
            month = int(exam_date[-2:])
            if month > 8:
                exam_date += "." + str(self.start_year)
            else:
                exam_date += "." + str(self.end_year)

            worktype = utils.get_worktype(exam_type)
            cur_color = ws.cell(row, col + 1).fill.start_color.index
            department_id = self.get_dep_id(cel_color=cur_color)
            disc_id = self.get_disc_id(lesson_name, department_id)

            # если перепутали название дисциплины, добавим такую
            if disc_id is None:
                disc_id = self.db.set_disc(
                    title="null",
                    shorttitle=lesson_name,
                    department_id=department_id,
                    varmask="null",
                )

            cur_date = datetime.strptime(exam_date, "%d.%m.%Y").date()

            # если вне рабочего промежутка, не пишем больше
            if cur_date > self.last_day:
                break
            # если не дошли до нужной даты
            if cur_date < self.first_day:
                continue

            params = {"day": cur_date, "semcode": self.semcode}
            day_id = self.db.get_id(table_name="sc_rasp18_days", params=params)

            if teacher is not None:
                params = {"fio": teacher}
                prep_id = self.db.get_id(table_name="sc_prep", params=params)
                # если перепутали имя препода, добавим такую
                if prep_id is None:
                    prep_id = self.db.set_prep(fio=teacher)

            # set_rasp18
            rasp18_id = self.db.set_rasp18(
                semcode=self.semcode,
                day_id=day_id,
                pair=cur_order,
                kind=0,
                worktype=worktype,
                disc_id=disc_id,
                timestart=str(time_start),
                timeend=str(time_end),
                prep_id=prep_id,
                room=room,
            )
            if rasp18_id is None:
                return

            # set_rasp18_groups
            self.db.set_rasp18_groups(
                rasp18_id=rasp18_id, group_id=group_id, subgroup=0
            )
            # set_rasp18_preps
            if prep_id is not None:
                self.db.set_rasp18_preps(rasp18_id=rasp18_id, prep_id=prep_id)
            # set_rasp18_rooms
            if room is not None:
                self.db.set_rasp18_rooms(rasp18_id=rasp18_id, room=room)

            # пропускаем уже разобранные ячейки экзамена
            # progress_bar.next(3)
            row += 3

        # progress_bar.finish()

    def get_lesson_parts(self, lesson_cell: str, is_magic: bool) -> LessonParts:
        """
        Разделяет пару на название дисциплины и доп информацию:
        подгруппа, недели, тип пары (лк, пр, лб)
        """
        lesson_parts = LessonParts()

        # получаем подгруппу
        lesson_parts.sub_group = self.get_subgroup(lesson_cell)

        # получаем тип пары (лк, пр, лб)
        lesson_type = utils.get_lesson_type(lesson_cell)
        lesson_parts.worktype = utils.get_worktype(lesson_type)
        # получаем данные по неделям
        weeks_parts = self.get_weeks_parts(lesson_cell, is_magic)
        lesson_parts.parity = weeks_parts.parity
        lesson_parts.weeks_list = weeks_parts.weeks_list
        lesson_parts.weeks_text = weeks_parts.weeks_text

        # удаляем лишнее и получаем название дисциплины
        lesson_parts.disc_name = utils.get_disc_name(lesson_cell, lesson_parts)

        return lesson_parts

    def get_subgroup(self, lesson: str) -> int:
        """Получить подгруппу"""
        subgroups = ListData.SUBGROUPS.value
        for sub_gr in subgroups:
            if sub_gr in lesson:
                return subgroups.index(sub_gr) + 1
        return 0

    # TODO упростить функцию
    def get_weeks_parts(self, lesson: LessonParts, is_magic: bool) -> WeeksParts:
        """Получить номера недель, когда будет пара"""
        last_week = WEEKS
        start_week = 1

        if is_magic:
            last_week = start_week + WEEKS
            start_week = self.db.get_week(self.magic_first_stud_date, self.semcode)

        all_weeks = list(range(start_week, last_week))

        # ищем кр.(кроме таких-то недель)
        except_weeks = re.match(utils.Patterns.EXCEPT_WEEKS, lesson)
        if except_weeks is not None:
            except_weeks = except_weeks.group(1).split(",")
            except_weeks = map(int, except_weeks)
            all_weeks = [
                w for w in range(start_week, last_week) if w not in except_weeks
            ]

        weeks_text = ", ".join(map(str, all_weeks))
        week_parts = WeeksParts()
        week_parts.weeks_list = all_weeks
        week_parts.weeks_text = weeks_text
        week_parts.parity = utils.get_week_parity(lesson)

        # генирируем массив всех недель
        if week_parts.parity != 0:
            parity_list = ListData.PARITY.value
            if start_week > week_parts.parity:
                is_even = start_week % 2
                if week_parts.parity % 2 != is_even:
                    start_week += 1
            else:
                start_week = week_parts.parity

            week_parts.weeks_list = list(range(start_week, last_week, 2))
            week_parts.weeks_text = parity_list[week_parts.parity - 1]

            # если была пометка кр. (кроме таких-то недель)
            if except_weeks is not None:
                week_parts.weeks_list = [
                    w for w in week_parts.weeks_list if w not in except_weeks
                ]
                weeks_text = ", ".join(map(str, week_parts.weeks_list))
            return week_parts

        # проверяем наличие недельного распределения
        substring = re.search(utils.Patterns.ONLY_STUD_WEEKS, lesson)
        if substring is not None and except_weeks is None:
            substring = substring.group().removesuffix("н")
            week_parts.weeks_text = substring
            weeks = substring.split(",")

            # если указан промежуток недель (5-8н)
            between_weeks = r"\d{1,2}-\d{1,2}"
            weeks_period = re.search(between_weeks, substring)
            if weeks_period is not None:
                start_week, last_week = weeks_period.group().split("-")
                weeks = list(range(int(start_week), int(last_week) + 1))

            week_parts.weeks_list = list(map(int, weeks))

        return week_parts

    def get_semcode(self, title: str) -> int:
        """Получение кода семестра по заголовку"""
        # парсим заголовок и получаем код семестра:
        # осень - 00, весна - 01

        # год в формате 2425, где 2024 - год начала учебного года, 2025 - год конца
        year_code = str(self.start_year)[-2:] + str(self.end_year)[-2:]

        semcode = "00"
        low_title = title.lower()
        is_spring = low_title.find(SPRING_SUBSTR)
        is_summer = low_title.find(SUMMER_SUBSTR)
        # если это весенний семестр, или летняя сессия, то четный семестр
        if is_spring != -1 or is_summer != -1:
            semcode = "01"
        semcode = str(year_code) + semcode

        return int(semcode)

    # TODO переименовать
    def fill_group_day_db(
        self,
        lesson_parts: LessonParts,
        teacher: str,
        order: int,
        weekday: str,
        group_id: int,
        room: str,
        department_id: int,
        is_magic: bool,
    ):
        """Заполнение таблиц по данным одного дня недели определенной группы"""
        # таблица дисциплин
        disc_id = self.db.set_disc(
            title="null",
            shorttitle=lesson_parts.disc_name,
            department_id=department_id,
            varmask="null",
        )

        # таблица расписания
        week_strs = ListData.WEEK_STRS.value
        weekday_num = week_strs.index(weekday)
        rasp7_id = self.db.set_rasp7(
            semcode=self.semcode,
            version=self.version,
            disc_id=disc_id,
            weekday=weekday_num,
            pair=order,
            weeksarray=lesson_parts.weeks_list,
            weekstext=lesson_parts.weeks_text,
        )

        # таблица связи пары
        self.db.set_rasp7_groups(
            rasp7_id=rasp7_id, group_id=group_id, sub_group=lesson_parts.sub_group
        )

        # распределение пар по преподавателям
        # disc_teachers = []
        # if teacher.find("\n") != -1:
        #     disc_teachers = teacher.split("\n")
        # else:
        #     disc_teachers.append(teacher)
        # # если преподавателей несколько, делаем несколько записей
        # prep_ids = []
        # # for t in disc_teachers:
        # #     # таблица преподавателей
        # #     prep_ids.append(self.db.set_prep(fio=t))

        # for prep_id in prep_ids:
        #     if prep_id is not None:
        prep_id = self.db.set_prep(fio=teacher)
        self.db.set_rasp7_preps(rasp7_id=rasp7_id, prep_id=prep_id)

        # заполнение аудиторий
        self.db.set_rasp7_rooms(rasp7_id=rasp7_id, room=room)

        ############################
        # заполнение таблиц rasp18 #
        ############################

        week_strs = ListData.WEEK_STRS.value

        day_order = (week_strs.index(weekday) - 1) % 7
        first_stud_day = datetime.strptime(self.start_date, "%Y-%m-%d").date()
        weekday_delta = abs(day_order - first_stud_day.weekday())

        weekday_num = week_strs.index(weekday)
        start_week = self.db.get_week(self.first_day, self.semcode)

        # TODO переписать, не нравится
        # магистры могут начинать учиться чуть позже
        if is_magic:
            start_week = self.db.get_week(self.magic_first_stud_date, self.semcode)

        params = {
            "semcode": self.semcode,
            "day": str(self.first_day),  # не влияет ни на что, перезапишется сразу
            "weekday": weekday_num,
            "week": start_week,
        }
        for week in lesson_parts.weeks_list:
            if week < start_week:
                continue
            # считаем дату пары
            cur_delta = weekday_delta + (week - 1) * 7
            cur_date = first_stud_day + timedelta(days=cur_delta)
            params["week"] = week
            params["day"] = str(cur_date)

            # если вне рабочего промежутка, не пишем больше
            if cur_date > self.last_day:
                break

            # смотрим id дня
            day_id = self.db.get_id(table_name="sc_rasp18_days", params=params)
            # время пары
            time_start = utils.get_time_by_order(order)
            time_end = utils.time_in_90_minutes(time_start)

            # если было указано какое-то особое время, то записываем
            if lesson_parts.timestart is not None:
                # если поставили НАЧАЛО пары пораньше
                time_start = deepcopy(lesson_parts.timestart)

                # если поставили КОНЕЦ пары пораньше
                h, m = lesson_parts.timeend.split(":")
                t_end_buf = time(hour=int(h), minute=int(m))
                time_end = min(t_end_buf, time_end)

            rasp18_id = self.db.set_rasp18(
                semcode=self.semcode,
                day_id=day_id,
                pair=order,
                kind=0,
                worktype=lesson_parts.worktype,
                disc_id=disc_id,
                timestart=str(time_start),
                timeend=str(time_end),
                prep_id=prep_id,
                room=room,
            )
            if rasp18_id is None:
                return

            self.db.set_rasp18_groups(rasp18_id, group_id, lesson_parts.sub_group)
            if prep_id is not None:
                self.db.set_rasp18_preps(rasp18_id, prep_id)
            if room is not None:
                self.db.set_rasp18_rooms(rasp18_id, room)

    def fill_rasp_title_parts(self, max_row: int, ws):
        """
        Заполнить по заголовоку расписания нужные для парсера данные:
        учебные года, версия, время года семестра

        return start_year, end_year, version
        """
        rasp_title = ""
        is_version = False
        for row in range(1, max_row):
            rasp_title = ws.cell(row, self.version_col).value
            if rasp_title is None or not isinstance(rasp_title, str):
                continue

            # если версия есть
            if rasp_title.find("версия") != -1:
                is_version = True
                break

        if not is_version:
            raise ValueError("Версия расписания не была найдена")

        self.version = utils.get_version(rasp_title)
        self.start_year, self.end_year = utils.get_stud_years(rasp_title)
        self.semcode = self.get_semcode(rasp_title)
        # TODO подумать и мб надо такое сделать
        # self.start_date, self.end_date = utils.get_stud_period(self.semcode)

    def get_disc_id(self, lesson_name: str, department_id: int) -> int:
        """Получить id дисциплины по названию"""
        params = {"shorttitle": lesson_name, "department_id": department_id}
        disc_id = self.db.get_id(table_name="sc_disc", params=params)
        return disc_id

    def get_dep_id(self, cel_color: int) -> int:
        """Получить по цвету ячейки id кафедры"""
        department_name = "другая"
        excel_color = Color(rgb=cel_color)

        if excel_color in CellColors.VEGA_DEP_COLORS:
            department_name = "ВЕГА"
        elif excel_color in CellColors.ONLY_VEGA_DEP_COLORS:
            department_name = "только для ВЕГИ"
        elif excel_color in CellColors.VM_DEP_COLORS:
            department_name = "ВМ"
        elif excel_color in CellColors.ONLY_VM_DEP_COLORS:
            department_name = "только для ВМ"
        elif excel_color in CellColors.OTHERS_DEP_COLORS:
            department_name = "другая"
        else:
            print(f"Новый цвет кафедры: {cel_color}")
            department_name = "другая"

        dep_params = {"title": department_name}
        department_id = self.db.get_id("sc_department", dep_params)
        return department_id


if __name__ == "__main__":
    parser = VegaRaspParser()
    parser.parse()
